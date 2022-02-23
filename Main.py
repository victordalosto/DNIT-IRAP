# ============================================================== #
# SCRIPT NAME : Main.py
# AUTHOR      : Victor Hugo Dalosto de Oliveira
# EMAIL       : victordalosto@gmail.com
# DESCRIPTION : Script that captures images from video and
# attributes to km in the road to be decodified via IRAP.
# ============================================================== #

# Imports that are used in the file. Some may need to be installed manually
import os
import csv
import shutil
import glob
import openpyxl
import xml.etree.ElementTree as ET
from datetime import datetime

# ============================================================== #
# Gets path of the current file
pathMain = os.getcwd()
# pathData is the Local Network where all files are located
pathData = "//10.100.10.219/Videos/Recebidos/"
# pathOutput is where the output will be stored
pathOutput = "E:/"
# ============================================================== #

# ============================================================== #
# =====================  F U N C T I O N S ===================== #
# ============================================================== #

# The folder where the output files of this script will be temporarily saved
pathStorage = os.path.join(pathMain, "Files")

# Final Output path folder where files will be exported
pathOutput = os.path.join(pathOutput, "OUTPUT")

# Path of the standard Excel input that is stored in each Road SNV folder
pathInput = os.path.join(pathMain, "lib", "model_input.xlsx")

# Path containing the CSV file with all Roads addresses in the local network
pathResume = os.path.join(pathMain, "lib", "Resumo.csv")

# Path containing the FFMPEG (C++ script) to Generate images from .mp4
# It may be needed to extract "FFMPEG" ../lib/ folder
pathFFMPEG = os.path.join(pathMain, "lib", "ffmpeg", 'bin')

# List of Roads SNVs
listSNVs = [[], []]

# Create an output log file using the time and date
dateTime = datetime.now()
logName = "LOG_" + dateTime.strftime('%d_%m_%Y_%H_%M_%S')+".txt"
pathReportLog = os.path.join(pathMain, "lib", logName)


# Function that creates the Main Folders for the Script to work
def createMainFolder():
    # Create the 'Files' Folder to store temporarily the files
    if (os.path.exists(pathStorage) is False):
        os.mkdir(pathStorage)
    # Creathe the final Output folder to store the files after verifications
    if (os.path.exists(pathOutput) is False):
        os.mkdir(pathOutput)
    # Create the library folder "..dir/lib/"
    if (os.path.exists(os.path.join(pathMain, "lib")) is False):
        os.mkdir(os.path.join(pathMain, "lib"))
    # Check if the CSV file with all Road SNV addresses exists
    if (os.path.isfile(pathResume) is False):
        print("Could not find the Input csv with Local Network addresses")
        exit()
    # Create the standard excel input used in the IRAP Codification
    if (os.path.isfile(pathInput) is False):
        wb = openpyxl.Workbook()  # Create sheet
        ws = wb['Sheet']  # Name of sheet
        ws.title = "_"  # Change name o sheet
        list1 = ["Rodovia", "Pista", "Faixa", "Início", "Fim", "Data"]
        list2 = ["Início", "Fim", "cam1 ini", "cam1 fim", "cam2 ini",
                 "cam2 fim", "TRR", "O", "P", "E", "Ex", "D", "R", "FI",
                 "J", "JE", "TB", "TBE", "TTC", "TTL", "TLC", "TLL", "ALP",
                 "ALC", "ATP", "ATC", "DG", "Observação", "Latitude",
                 "Longitude"]
        list3 = [8.285, 17.285, 8.42578, 9, 9, 9, 5, 5.570, 6.1406, 6.1406,
                 3.570, 2.7109, 2.7109, 3, 2.42578, 3.5703, 3.7109, 5,
                 4.85546875, 4.570, 4.855, 4.570, 5, 5, 5, 5, 4.1406,
                 11.42578125, 12.140625, 12.140625]
        for i in range(len(list1)):
            ws.cell(row=i+1, column=1).value = list1[i]
            ws.cell(row=i+1, column=1).font = openpyxl.styles.Font(bold=True)
        for i in range(len(list2)):
            ws.cell(row=8, column=i+1).value = list2[i]
            ws.cell(row=8, column=i+1).font = openpyxl.styles.Font(bold=True)
            style = openpyxl.styles.Alignment(horizontal='center')
            ws.cell(row=8, column=i+1).alignment = style
            column = openpyxl.utils.get_column_letter(i+1)
            ws.column_dimensions[column].width = list3[i]
        ws.cell(row=6, column=2).number_format = "DD/MM/YYYY"
        wb.save(filename=(pathInput))
        wb.close()
    # Create a report log with the status of all the imports
    if (os.path.isfile(pathReportLog) is False):
        reportLog = open(pathReportLog, "w")
    else:
        reportLog = open(pathReportLog, "r+")
        reportLog.truncate(0)
    reportLog.writelines("Log: " + dateTime.strftime("%d/%m/%Y %H:%M:%S\n\n"))
    reportLog.writelines("####### FAILED SNV IMPORT #######\n\n")
    reportLog.writelines("####### SNV IMPORT ERROR ########\n\n")
    reportLog.writelines("####### SUCESS SNV IMPORT #######\n\n")
    reportLog.close()


# Function that updates the Road SNV List
def updateList(nameFolder, MSG, typeMSG):
    listSNVs[0].append(nameFolder)
    listSNVs[1].append(MSG)
    if (typeMSG != ''):
        updateLog(typeMSG)


# function that update the Report Log txt file with all the importation status
def updateLog(condition):
    match condition:
        case "fail":
            conditionLog = "####### FAILED SNV IMPORT #######"
        case "problem":
            conditionLog = "####### SNV IMPORT ERROR ########"
        case "sucess":
            conditionLog = "####### SUCESS SNV IMPORT #######"
    list_of_lines = (open(pathReportLog, "r")).readlines()
    for i, line in enumerate(open(pathReportLog, "r")):
        if conditionLog in line:
            list_of_lines.insert(i+1, (listSNVs[0][-1])+" - "+listSNVs[1][-1])
            list_of_lines.insert(i+2, "\n")
            reportLog = open(pathReportLog, "w")
            reportLog.writelines(list_of_lines)
            reportLog.close()
            break


# Function that creates a list of all files that are already correctly imported
def excludeImportedSNV():
    rootdir = os.walk(pathOutput)
    for root, dir, file in rootdir:
        for nameSNV in file:
            if nameSNV.endswith('.xlsx'):
                nameSNV = nameSNV.replace(".xlsx", '')
                updateList(nameSNV, '', '')


# Function that stop the script if Disk Storage is full
def getDiskSpace(nameFolder):
    HD_total, HD_used, HD_freeStorage = shutil.disk_usage(pathStorage)
    HD_total, HD_used, HD_freeOutput = shutil.disk_usage(pathOutput)
    conv = 1/(1024*1024*1024)
    if HD_freeStorage*conv <= 35 or HD_freeOutput*conv <= 35:
        updateList(nameFolder, "Not enought disk space", "fail")
        print("\n\nProgram terminated due to not enought Disk Space\n\n")
        exit()


# Set the path of folder and files of the current selected Road SNV
def pathSet(nameFolder, addressSNV):
    pathFolder = os.path.join(pathStorage, nameFolder)
    pathImg = os.path.join(pathFolder, "Cam 1")
    pathSNV = os.path.join(pathData, addressSNV)
    pathXML = os.path.join(pathSNV, "LogsTrecho.xml")
    pathVideo = os.path.join(pathSNV, "videos", "camera1")
    pathExcel = os.path.join(pathFolder, (nameFolder + ".xlsx"))
    for file in glob.glob(os.path.join(pathVideo, "*.mp4")):
        pathVideo = os.path.join(pathVideo, file)  # path to Video
        break  # Prevents selecting more than one video
    return pathFolder, pathImg, pathXML, pathVideo, pathExcel


# Function that checks if selected Road SNV is able to be exported
def firstCheck(nameFolder, pathFolder, pathXML, pathVideo):
    if (os.path.isdir(pathFolder) is True):  # Folder already exists
        updateList(nameFolder, "Folder already exists", "fail")
        # print(listSNVs[1][-1] + ": " +listSNVs[0][-1])
    if (os.path.isfile(pathXML) is False):  # Road SNV doesn't have a Log.xml
        updateList(nameFolder, "Could not find Log.xml", "fail")
        # print(listSNVs[1][-1] + ": " +listSNVs[0][-1])
    if (pathVideo.endswith('.mp4') is False):  # Check if SNV has a valid .mp4
        updateList(nameFolder, "Could not find Video", "fail")
        # print(listSNVs[1][-1] + ": " +listSNVs[0][-1])


# Function that creates folders according to the input structure
def createSNVFolder(nameFolder, pathExcel):
    os.mkdir(os.path.join(pathStorage, nameFolder))
    os.mkdir(os.path.join(pathStorage, nameFolder, "Cam 1"))
    os.mkdir(os.path.join(pathStorage, nameFolder, "Cam 2"))
    shutil.copy2(pathInput, pathExcel)


# This function obtains the values of Odometer, frontal video timing,
# and geographical coordinates from LogsTrecho.xml for each Road SNV
def getAllOdometerValues(pathXML):
    root = ET.parse(pathXML).getroot()  # Open the Log.xml file
    # Create array containing for each Road SNV, all values of:
    # [[Odometer], [videoTiming], [latitude], [longitude]]
    odomArray = [[], [], [], []]
    for i in range(len(root[0])):
        try:
            #  Get the Odometer value in km
            odometer = float((root[0][i].attrib).get('Odometro'))
            #  Get time of video in second
            videoTiming = float((root[0][i][0].attrib).get('Frente'))
            #  Get the geographic longitude (X)
            longitude = float((root[0][i][1].attrib).get('X'))
            #  Get the geographic latitude (Y)
            latitude = float((root[0][i][1].attrib).get('Y'))
            # If script haven't found a error, attributes to odomArray[]
            odomArray[0].append(odometer)
            odomArray[1].append(videoTiming)
            odomArray[2].append(longitude)
            odomArray[3].append(latitude)
        except (TypeError, ValueError):
            pass
    return odomArray


# Loop that gets the closest odometer value to the IRAP input format
# spaced in 20-20m, and saves all its Data to select images and export
def filterOdometerIRAP(Array):
    imageArray = [[], [], [], []]
    size = len(Array[0])
    if ((size > 0) and (len(Array[1]) > 0)):
        spacing = 20  # Irap spacing = 20-20meters
        for odometer in range(0, int(round(max(Array[0]), 0)), spacing):
            # find closest index to odometer
            index = min(range(size), key=lambda x: abs(Array[0][x]-odometer))
            imageArray[0].append(odometer)  # New odometer value
            imageArray[1].append(Array[1][index])  # Video timming
            imageArray[2].append(Array[2][index])  # geograph longitude (x)
            imageArray[3].append(Array[3][index])  # geograph latitude  (y)
        # Include the last value if its already not in the list
        if int(round(max(Array[0]), 0)) not in imageArray[0]:
            odometer = int(round(max(Array[0]), 0))
            index = min(range(size), key=lambda x: abs(Array[0][x]-odometer))
            imageArray[0].append(odometer)  # New odometer value
            imageArray[1].append(Array[1][index])  # Video timming
            imageArray[2].append(Array[2][index])  # geograph longitude (x)
            imageArray[3].append(Array[3][index])  # geograph latitude  (y)
    return imageArray


# FFMPEG in the library that generates images in the "..dir/Cam 1/" folder.
# The script names the images as XXXXdum.png that will then be deleted
def createImages(pathVideo, pathImg):
    timing = 1  # number of frames to be used for each second
    input = '"'+pathVideo+'"'
    output = " "+'"'+pathImg+"/%ddum.png"+'"'
    os.system("ffmpeg -i " + input + " -y -r " + str(timing) + output)
    # To generate images with lower quality -> change png to jpeg)


# Function that selects from all the generated images,
# only the ones that are validated to IRAP spacing, deleting the remaining.
def deleteImages(pathImg, imageArray):
    if imageArray[0][0] <= 1.5 and imageArray[0][1] > 1.5:
        try:
            input = os.path.join(pathImg, "1dum.png")
            output = os.path.join(pathImg, "0000000000.png")
            os.rename(input, output)
        except FileNotFoundError:
            pass
    inicio = 0
    if os.path.isfile(os.path.join(pathImg, "0000000000.png")):
        inicio = min(1, len(imageArray[0]))
    for i in range(inicio, len(imageArray[0])):  # Rename all images in Folder
        try:
            timePhoto = str(int(round(imageArray[1][i], 0))) + "dum.png"
            odometer = str(format(imageArray[0][i], '010d')) + ".png"
            input = os.path.join(pathImg, timePhoto)
            output = os.path.join(pathImg, odometer)
            os.rename(input, output)
        except FileNotFoundError:
            pass
    # Delete all images that don't correspond to the IRAP spacing
    for file in os.listdir(pathImg):
        if file.endswith("dum.png"):
            if os.path.isfile(os.path.join(pathImg, file)):
                os.remove(os.path.join(pathImg, file))


# Function that gets the all the coordinates and put in the Excel Output
def updateExcelInput(nameFolder, imageArray, direction, pathExcel):
    wb = openpyxl.load_workbook(pathExcel)
    ws = wb.active
    ws.cell(row=1, column=2).value = nameFolder  # Saves road name in (1,1)
    ws.cell(row=2, column=2).value = direction   # Saves lane direction (2,1)
    ws.cell(row=3, column=2).value = 2  # Number of lanes - Users input
    ws.cell(row=4, column=2).value = 0  # Initial kilometer of Road SNV
    # Get the highest odometer value = final kilometer
    ws.cell(row=5, column=2).value = imageArray[0][len(imageArray[0])-1]/1000
    ws.cell(row=6, column=2).value = '=today()'  # Get today's value in Excel
    for i in range(len(imageArray[0])-1):
        rowCell = i + 9  # Initial Cell in Excel Input = 9
        ws.cell(row=rowCell, column=1).value = imageArray[0][i]/1000
        ws.cell(row=rowCell, column=2).value = (imageArray[0][i]+20)/1000
        ws.cell(row=rowCell, column=3).value = imageArray[0][i]
        ws.cell(row=rowCell, column=4).value = imageArray[0][i]+20
        ws.cell(row=rowCell, column=29).value = imageArray[3][i]
        ws.cell(row=rowCell, column=30).value = imageArray[2][i]
    lastCell = len(imageArray[0]) - 1 + 9 - 1
    ws.cell(row=lastCell, column=2).value = (imageArray[0][-1])/1000
    ws.cell(row=lastCell, column=4).value = (imageArray[0][-1])
    wb.save(filename=pathExcel)
    wb.close()


# Function that checks if the information inside the output folder is correct.
# Compare extensions, coordinates, and photos with the information in CSV
def finalCheck(nameFolder, pathImg, array, extensionCSV):
    # extensionCSV = input according to CSV File
    # Extension according to number of images
    extensionPhoto = len(os.listdir(pathImg))*20/1000
    # Extension according to data in log
    extensionLog = array[0][-1]/1000
    # Estimated precision for remove list from import
    precision = 0.2
    if extensionPhoto == 0:
        updateList(nameFolder, "0 Images generated", "problem")
    if extensionLog == 0:
        updateList(nameFolder, "Log = 0 km", "problem")
    if (not array[0]) and (not array[1]):
        updateList(nameFolder, "Couldn't obtain values from log", "problem")
    if abs(extensionCSV - extensionPhoto) > precision:
        eCSV = str(round(extensionCSV, 2))
        ePhoto = str(round(extensionPhoto, 2))
        MSG = "SNV (" + eCSV + ")km =/= Photo Extension (" + ePhoto + ")km"
        updateList(nameFolder, MSG, "problem")
    if abs(extensionCSV - extensionLog) > precision:
        eCSV = str(round(extensionCSV, 2))
        eLOG = str(round(extensionLog, 2))
        MSG = "SNV (" + eCSV + ")km =/= Log Extension (" + eLOG + ")km"
        updateList(nameFolder, MSG, "problem")
    corruptedImages = 0
    for root, dir, images in os.walk(pathImg):
        for file in images:
            if os.path.exists(os.path.join(pathImg, file)):
                if os.path.getsize(os.path.join(pathImg, file)) == 0:
                    corruptedImages += 1
        if corruptedImages > 2:
            updateList(nameFolder, "Problem - Corrupted Images", "problem")


# Function that create folders for states-UF and BR road names
def createStateBRFolders(nameBR, nameUF):
    pathUF = os.path.join(pathOutput, nameUF)
    pathBR = os.path.join(pathUF, nameBR)
    if (os.path.exists(pathUF) is False):
        os.mkdir(os.path.join(pathOutput, nameUF))
    if (os.path.exists(pathBR) is False):
        os.mkdir(os.path.join(pathUF, nameBR))


# Function that move SNV folder to the desired OUTPUT after check
def moveFolders(original, target):
    shutil.move(original, target)


# Function that loop through all functions
def Main():
    createMainFolder()
    with open(pathResume) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')
        next(csv_reader)  # jump header of CSV file
        os.chdir(pathFFMPEG)  # Folder where is stored FFMPEG script
        excludeImportedSNV()
        for row in csv_reader:
            if (row[0] != ''):
                nameFolder = row[1] + "_" + row[0]  # Generating road name
                addressSNV = row[8]  # Address of Road SNV in the Network
                getDiskSpace(nameFolder)
                allPaths = pathSet(nameFolder, addressSNV)
                pathFolder, pathImg, pathXML, pathVideo, pathExcel = allPaths
                firstCheck(nameFolder, pathFolder, pathXML, pathVideo)
                if nameFolder not in listSNVs[0]:
                    createSNVFolder(nameFolder, pathExcel)
                    array1 = getAllOdometerValues(pathXML)
                    array = filterOdometerIRAP(array1)
                    createImages(pathVideo, pathImg)
                    deleteImages(pathImg, array)
                    if (float(row[4]) < float(row[5])):
                        direction = "Crescente"
                    else:
                        direction = "Decrescente"
                    updateExcelInput(nameFolder, array, direction, pathExcel)
                    finalCheck(nameFolder, pathImg, array, float(row[6]))
                    if nameFolder not in listSNVs[0]:
                        createStateBRFolders("BR-" + row[2], row[3])
                        File = os.path.join(row[3], "BR-" + row[2], nameFolder)
                        moveFolders(pathFolder, os.path.join(pathOutput, File))
                        updateList(nameFolder, "Sucessfull imported", "sucess")


Main()
